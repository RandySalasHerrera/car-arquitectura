import os
from coursescraped import models as models_course_scraped
from student import models as models_student
from teacher import models as models_teacher
from course import models as models_course
from qualification import models as models_qualification
from openpyxl import load_workbook
from django.template.loader import get_template
from django.core.mail import send_mail
import threading


def createUser(data):
    try:
        estudiantes_creados = 0
        estudiantes_error = 0
        for fila in data.iter_rows(min_row=2, values_only=True):
            name = fila[0].lower()
            sex  = fila[1]
            age  = fila[2]
            address = fila[3]
            email = fila[4]

            if name and isinstance(name, str):
                if sex and isinstance(sex, str) and age and isinstance(age, int) and address and isinstance(address, str) and email and isinstance(email, str):
                    if not models_student.Student.objects.filter(name=name).exists():
                        models_student.Student.objects.create(
                            name=name,
                            sex=sex,
                            age=age,
                            address=address,
                            email=email
                        )
                        estudiantes_creados += 1
                    else:
                        estudiantes_error += 1
                else:
                    estudiantes_error += 1
            else:
                estudiantes_error += 1
                    
        return estudiantes_creados, estudiantes_error
    except Exception as e:
        estudiantes_error += 1
        return estudiantes_creados, estudiantes_error


def createTeacher(data):
    try:
        teacher_creados = 0
        teacher_error = 0
        for fila in data.iter_rows(min_row=2, values_only=True):
            name = fila[0].lower()
            sex = fila[1]
            age = fila[2]
            address = fila[3]
            email = fila[4]

            if name and isinstance(name, str):
                if sex and isinstance(sex, str) and age and isinstance(age, int) and address and isinstance(address, str) and email and isinstance(email, str):
                    if not models_teacher.Teacher.objects.filter(name=name).exists():
                        models_teacher.Teacher.objects.create(
                            name=name,
                            sex=sex,
                            age=age,
                            address=address,
                            email=email
                        )
                        teacher_creados += 1
                    else:
                        teacher_error += 1
                else:
                    teacher_error += 1
            else:
                teacher_error += 1
                    
        return teacher_creados, teacher_error
    except Exception as e:
        teacher_error += 1
        return teacher_creados, teacher_error


def createCourse(data):
    try:
        course_creados = 0
        course_error = 0
        for fila in data.iter_rows(min_row=2, values_only=True):
            course_scraped_name = fila[0].lower()
            teacher_name = fila[1].lower()
            student_name = fila[2].lower()

            if course_scraped_name and isinstance(course_scraped_name, str) and teacher_name and isinstance(teacher_name, str) and student_name and isinstance(student_name, str) :

                if not models_course_scraped.CursoScraped.objects.filter(name=course_scraped_name).exists():
                    course_scraped = models_course_scraped.CursoScraped.objects.create(name=course_scraped_name)

                course_scraped = models_course_scraped.CursoScraped.objects.filter(name=course_scraped_name).first()

                if not models_teacher.Teacher.objects.filter(name=teacher_name).exists():
                    models_teacher.Teacher.objects.create(name=teacher_name)

                teacher = models_teacher.Teacher.objects.filter(name=teacher_name).first()

                if not models_student.Student.objects.filter(name=student_name).exists():
                    models_student.Student.objects.create(name=student_name)

                student = models_student.Student.objects.filter(name=student_name).first()

                if not models_course.Course.objects.filter(course_scraped=course_scraped).exists():
                    course = models_course.Course.objects.create(
                        course_scraped=course_scraped,
                        teacher=teacher,
                    )
                    course.students.add(student)
                    course_creados += 1
                else:
                    course = models_course.Course.objects.filter(course_scraped=course_scraped).first()
                    course.students.add(student)
                    course_creados += 1
            
            else:
                course_error += 1
                    
        return course_creados, course_error
    except Exception as e:
        course_error += 1
        return course_creados, course_error



def createQualification(data):
    try:
        quialification_creados = 0
        quialification_error = 0
        for fila in data.iter_rows(min_row=2, values_only=True):
            course_name = fila[0].lower()
            student = fila[1].lower()
            quialification = fila[2]

            if course_name and isinstance(course_name, str):
                if student and isinstance(student, str) and quialification and isinstance(quialification, int) :
                    if models_course.Course.objects.filter(course_scraped__name=course_name, students__name=student).exists():

                            get_course = models_course.Course.objects.filter(course_scraped__name=course_name, students__name=student).first()
                            get_student = models_student.Student.objects.filter(name=student).first()

                            if not models_qualification.Qualification.objects.filter(course=get_course, student=get_student).exists():
                                models_qualification.Qualification.objects.create(
                                    course=get_course,
                                    student=get_student,
                                    qualification=quialification
                                )

                                quialification_creados += 1

                    else:
                        quialification_error += 1
                else:
                    quialification_error += 1
            else:
                quialification_error += 1
                    
        return quialification_creados, quialification_error
    except Exception as e:
        teacher_error += 1
        return quialification_creados, quialification_error



def emailNotification(data):
    context = data
    template = get_template("notificacion.html")
    context = template.render(context)

    send_mail(
        "Notificacion Prueba De Arquitectura",
        "",
        "dashboardbot@lsv-tech.com",
        # [os.environ.get("EMAIL_NOTIFICATION", "")],
        ["salasrandy89@gmail.com"],
        fail_silently=True,
        html_message=context,
    )

def threading_emailNotification(data):
    thread = threading.Thread(target=emailNotification, args=(data,), daemon=False)
    thread.start()


def importFile(excel_upload):
    wb = load_workbook(excel_upload.file)
    nombres_hojas = ['Hoja1', 'Hoja2', 'Hoja3', 'Hoja4']
    for nombre_hoja in nombres_hojas:

        hoja = wb[nombre_hoja]

        if nombre_hoja == 'Hoja1':
            try:
                student = createUser(hoja)
                estudiantes_creados, estudiantes_error = student
                print(estudiantes_creados, estudiantes_error, 'student')
            except:
                pass

        if nombre_hoja == 'Hoja2':
            try:
                teacher = createTeacher(hoja)
                teacher_creados, teacher_error = teacher
                print(teacher_creados, teacher_error, 'teacher')
            except:
                pass

        if nombre_hoja == 'Hoja3':
            try:
                course = createCourse(hoja)
                course_creados, course_error = course
                print(course_creados, course_error, 'course')
            except:
                pass

        if nombre_hoja == 'Hoja4':
            try:
                qualification = createQualification(hoja)
                qualification_creados, qualification_error = qualification
                print(qualification_creados, qualification_error, 'qualification')
            except:
                pass

    wb.close()
    data = {
        'estudiantes_creados': estudiantes_creados,
        'estudiantes_error': estudiantes_error,
        'teacher_error': teacher_error,
        'teacher_creados': teacher_creados,
        'course_creados': course_creados,
        'course_error': course_error,
        'qualification_creados': qualification_creados,
        'qualification_error': qualification_error,
    }

    threading_emailNotification(data)


